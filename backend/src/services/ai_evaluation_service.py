import csv
import json
from collections import Counter
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.domain.ai_models import AIAssessmentSuggestion
from src.domain.enums import AISuggestionStatus, ApplicabilityStatus, AssessmentVerdict
from src.services.ai_assessment_service import AIAssessmentCandidate


MANUAL_REVIEW_FIELDS = (
    "requirement_id",
    "standard_verified",
    "manual_applicability",
    "suggested_verdict",
    "evidence_validity",
    "correct_pdf_pages",
    "rationale_correct",
    "missing_items_correct",
    "review_complete",
)
V1_APPLICABILITY_EXCEPTION_IDS = frozenset({"GRI 2-30-b"})
FINAL_ADJUDICATION_FIELDS = (
    "requirement_id",
    "final_applicability",
    "final_verdict",
    "final_pdf_pages",
    "final_evidence_validity",
    "final_rationale",
    "final_missing_items",
    "decision_basis",
    "adjudicator_role",
    "adjudication_version",
)
FINAL_EVIDENCE_VALIDITIES = frozenset(
    {"valid", "partially_valid", "invalid", "none"}
)


def load_adjudication_pending_ids(path: Path) -> set[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        if "requirement_id" not in (reader.fieldnames or []):
            raise ValueError("adjudication recommendations missing requirement_id")
        requirement_ids = {
            _text(row.get("requirement_id"))
            for row in reader
            if _text(row.get("requirement_id"))
        }
    if not requirement_ids:
        raise ValueError("adjudication recommendations contain no requirement_id")
    return requirement_ids


@dataclass(frozen=True)
class ManualReviewRecord:
    requirement_id: str
    standard_verified: str
    manual_applicability: str
    suggested_verdict: AssessmentVerdict | None
    evidence_validity: str
    correct_pdf_pages: list[int]
    rationale_correct: str
    missing_items_correct: str
    review_complete: str
    is_applicability_exception: bool = False


@dataclass(frozen=True)
class ManualReviewBaseline:
    report_id: str
    run_id: str
    records: list[ManualReviewRecord]

    @property
    def by_id(self) -> dict[str, ManualReviewRecord]:
        return {record.requirement_id: record for record in self.records}


@dataclass(frozen=True)
class FinalAdjudicationRecord:
    requirement_id: str
    final_applicability: ApplicabilityStatus
    final_verdict: AssessmentVerdict
    final_pdf_pages: list[int]
    final_evidence_validity: str
    final_rationale: str
    final_missing_items: list[str]
    decision_basis: str
    adjudicator_role: str
    adjudication_version: str


@dataclass(frozen=True)
class AIEvaluationCase:
    manual: ManualReviewRecord
    candidate: AIAssessmentCandidate


@dataclass(frozen=True)
class AIEvaluationResult:
    rows: list[dict[str, Any]]
    summary: dict[str, Any]


def load_manual_review_baseline(
    path: Path,
    *,
    expected_count: int = 225,
) -> ManualReviewBaseline:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        report_id, run_id = _read_dataset_identity(workbook)
        worksheet, header_row, headers = _find_review_sheet(workbook)
        missing = [field for field in MANUAL_REVIEW_FIELDS if field not in headers]
        if missing:
            raise ValueError(f"manual review workbook missing columns: {missing}")

        records: list[ManualReviewRecord] = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            raw = {name: values[index] if index < len(values) else None for name, index in headers.items()}
            requirement_id = _text(raw["requirement_id"])
            if not requirement_id:
                continue
            standard_verified = _text(raw["standard_verified"]).lower()
            review_complete = _text(raw["review_complete"]).lower()
            manual_applicability = _text(raw["manual_applicability"])
            is_exception = (
                requirement_id in V1_APPLICABILITY_EXCEPTION_IDS
                and standard_verified == "yes"
                and manual_applicability == "not_applicable_confirmed"
                and not _text(raw["suggested_verdict"])
            )
            if standard_verified != "yes" or (
                review_complete != "complete" and not is_exception
            ):
                continue
            verdict_text = _text(raw["suggested_verdict"])
            records.append(
                ManualReviewRecord(
                    requirement_id=requirement_id,
                    standard_verified=standard_verified,
                    manual_applicability=manual_applicability,
                    suggested_verdict=(
                        AssessmentVerdict(verdict_text) if verdict_text else None
                    ),
                    evidence_validity=_text(raw["evidence_validity"]),
                    correct_pdf_pages=_parse_pages(raw["correct_pdf_pages"]),
                    rationale_correct=_text(raw["rationale_correct"]),
                    missing_items_correct=_text(raw["missing_items_correct"]),
                    review_complete=review_complete,
                    is_applicability_exception=is_exception,
                )
            )
        if len(records) != expected_count:
            raise ValueError(
                f"manual review selection count mismatch: expected {expected_count}, got {len(records)}"
            )
        requirement_ids = [record.requirement_id for record in records]
        if len(set(requirement_ids)) != len(requirement_ids):
            raise ValueError("manual review selection contains duplicate requirement_id")
        return ManualReviewBaseline(
            report_id=report_id,
            run_id=run_id,
            records=records,
        )
    finally:
        workbook.close()


def load_final_adjudications(
    path: Path,
    *,
    valid_requirement_ids: set[str] | None = None,
    expected_count: int | None = 16,
) -> dict[str, FinalAdjudicationRecord]:
    records: dict[str, FinalAdjudicationRecord] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        missing_fields = sorted(
            set(FINAL_ADJUDICATION_FIELDS).difference(reader.fieldnames or [])
        )
        if missing_fields:
            raise ValueError(
                f"final adjudication CSV missing columns: {missing_fields}"
            )
        for row in reader:
            requirement_id = _text(row.get("requirement_id"))
            if not requirement_id:
                continue
            if requirement_id in records:
                raise ValueError(
                    f"duplicate final adjudication: {requirement_id}"
                )
            if (
                valid_requirement_ids is not None
                and requirement_id not in valid_requirement_ids
            ):
                raise ValueError(
                    "final adjudication is not present in GRI scope: "
                    f"{requirement_id}"
                )
            try:
                final_applicability = ApplicabilityStatus(
                    _text(row.get("final_applicability"))
                )
                final_verdict = AssessmentVerdict(
                    _text(row.get("final_verdict"))
                )
            except ValueError as exc:
                raise ValueError(
                    f"invalid final adjudication enum: {requirement_id}"
                ) from exc
            evidence_validity = _text(row.get("final_evidence_validity"))
            if evidence_validity not in FINAL_EVIDENCE_VALIDITIES:
                raise ValueError(
                    f"invalid final evidence validity: {requirement_id}"
                )
            final_rationale = _text(row.get("final_rationale"))
            decision_basis = _text(row.get("decision_basis"))
            adjudicator_role = _text(row.get("adjudicator_role"))
            adjudication_version = _text(row.get("adjudication_version"))
            if not all(
                (
                    final_rationale,
                    decision_basis,
                    adjudicator_role,
                    adjudication_version,
                )
            ):
                raise ValueError(
                    f"incomplete final adjudication metadata: {requirement_id}"
                )
            records[requirement_id] = FinalAdjudicationRecord(
                requirement_id=requirement_id,
                final_applicability=final_applicability,
                final_verdict=final_verdict,
                final_pdf_pages=_parse_json_pages(
                    row.get("final_pdf_pages"),
                    requirement_id=requirement_id,
                ),
                final_evidence_validity=evidence_validity,
                final_rationale=final_rationale,
                final_missing_items=_parse_json_strings(
                    row.get("final_missing_items"),
                    requirement_id=requirement_id,
                ),
                decision_basis=decision_basis,
                adjudicator_role=adjudicator_role,
                adjudication_version=adjudication_version,
            )
    if expected_count is not None and len(records) != expected_count:
        raise ValueError(
            "final adjudication count mismatch: "
            f"expected {expected_count}, got {len(records)}"
        )
    return records


def apply_final_adjudications(
    baseline: ManualReviewBaseline,
    adjudications: dict[str, FinalAdjudicationRecord],
) -> ManualReviewBaseline:
    baseline_by_id = baseline.by_id
    missing_ids = sorted(set(adjudications).difference(baseline_by_id))
    if missing_ids:
        raise ValueError(
            "final adjudication is not present in manual baseline: "
            + ", ".join(missing_ids)
        )
    records = [
        (
            replace(
                record,
                manual_applicability=adjudications[
                    record.requirement_id
                ].final_applicability.value,
                suggested_verdict=adjudications[
                    record.requirement_id
                ].final_verdict,
                evidence_validity=adjudications[
                    record.requirement_id
                ].final_evidence_validity,
                correct_pdf_pages=list(
                    adjudications[record.requirement_id].final_pdf_pages
                ),
                rationale_correct="yes",
                missing_items_correct="yes",
                review_complete="complete",
                is_applicability_exception=False,
            )
            if record.requirement_id in adjudications
            else record
        )
        for record in baseline.records
    ]
    return ManualReviewBaseline(
        report_id=baseline.report_id,
        run_id=baseline.run_id,
        records=records,
    )


def evaluate_ai_suggestions(
    cases: list[AIEvaluationCase],
    suggestions: list[AIAssessmentSuggestion],
    *,
    adjudication_pending_ids: set[str] | None = None,
) -> AIEvaluationResult:
    adjudication_pending_ids = adjudication_pending_ids or set()
    suggestion_by_assessment = {
        suggestion.assessment_id: suggestion for suggestion in suggestions
    }
    rows: list[dict[str, Any]] = []
    exact_count = 0
    false_disclosed_count = 0
    unsupported_count = 0
    wrong_page_count = 0
    schema_failure_count = 0
    model_failure_count = 0
    disagreement_count = 0
    manual_page_disagreement_count = 0
    verdict_evaluable_count = 0
    cross_distribution: Counter[str] = Counter()
    all_rows_false_disclosed_count = 0
    all_rows_wrong_page_count = 0

    for case in cases:
        manual = case.manual
        is_adjudication_pending = manual.requirement_id in adjudication_pending_ids
        candidate = case.candidate
        suggestion = suggestion_by_assessment.get(candidate.assessment.assessment_id)
        manual_label = (
            manual.suggested_verdict.value
            if manual.suggested_verdict
            else "applicability_exception"
        )
        cross_distribution[f"{manual_label}|{candidate.review_priority.value}"] += 1
        if manual.suggested_verdict is not None:
            verdict_evaluable_count += 1

        status = suggestion.status.value if suggestion else "missing"
        ai_verdict = (
            suggestion.suggested_verdict
            if suggestion and suggestion.status is AISuggestionStatus.SUCCEEDED
            else None
        )
        exact = bool(
            manual.suggested_verdict is not None
            and ai_verdict is manual.suggested_verdict
        )
        exact_count += int(exact)
        false_disclosed = bool(
            ai_verdict is AssessmentVerdict.DISCLOSED
            and manual.suggested_verdict is not AssessmentVerdict.DISCLOSED
        )
        all_rows_false_disclosed_count += int(false_disclosed)
        false_disclosed_count += int(false_disclosed and not is_adjudication_pending)

        guardrail_codes = suggestion.guardrail_codes if suggestion else []
        unsupported = "evidence_reference_out_of_scope" in guardrail_codes
        schema_failure = "response_schema_invalid" in guardrail_codes
        unsupported_count += int(unsupported)
        schema_failure_count += int(schema_failure)

        cited_pages = suggestion.evidence_pdf_pages if suggestion and ai_verdict else []
        correct_pages = set(manual.correct_pdf_pages)
        manual_page_disagreement = bool(
            ai_verdict is not None
            and ai_verdict is not AssessmentVerdict.UNKNOWN
            and cited_pages
            and any(page not in correct_pages for page in cited_pages)
        )
        manual_page_disagreement_count += int(manual_page_disagreement)
        wrong_page = bool(
            manual_page_disagreement
            and manual.suggested_verdict is ai_verdict
            and bool(correct_pages)
        )
        all_rows_wrong_page_count += int(wrong_page)
        wrong_page_count += int(wrong_page and not is_adjudication_pending)

        model_failure = bool(
            suggestion is None
            or (
                suggestion.status is AISuggestionStatus.FAILED
                and (
                    (suggestion.error_code or "").startswith("llm_")
                    or suggestion.error_code == "ai_service_unexpected_error"
                )
            )
        )
        model_failure_count += int(model_failure)
        disagreement = bool(
            ai_verdict is not None
            and ai_verdict is not candidate.assessment.verdict
        )
        disagreement_count += int(disagreement)

        rows.append(
            {
                "requirement_id": manual.requirement_id,
                "manual_applicability": manual.manual_applicability,
                "manual_verdict": manual.suggested_verdict.value if manual.suggested_verdict else "",
                "is_applicability_exception": manual.is_applicability_exception,
                "is_adjudication_pending": is_adjudication_pending,
                "manual_evidence_validity": manual.evidence_validity,
                "manual_correct_pdf_pages": manual.correct_pdf_pages,
                "rule_verdict": candidate.assessment.verdict.value,
                "review_priority": candidate.review_priority.value,
                "ai_status": status,
                "ai_suggested_verdict": ai_verdict.value if ai_verdict else "",
                "ai_evidence_pdf_pages": cited_pages,
                "ai_model": suggestion.model if suggestion else "",
                "ai_prompt_version": suggestion.prompt_version if suggestion else "",
                "guardrail_codes": guardrail_codes,
                "error_code": suggestion.error_code if suggestion else "missing_suggestion",
                "exact_verdict_agreement": exact,
                "false_disclosed": false_disclosed,
                "unsupported_evidence_reference": unsupported,
                "wrong_source_page": wrong_page,
                "manual_evidence_page_disagreement": manual_page_disagreement,
                "schema_failure": schema_failure,
                "model_failure": model_failure,
                "rules_ai_disagreement": disagreement,
            }
        )

    evaluated_count = len(cases)
    summary = {
        "evaluated_count": evaluated_count,
        "verdict_evaluable_count": verdict_evaluable_count,
        "applicability_exception_count": evaluated_count - verdict_evaluable_count,
        "exact_verdict_agreement_count": exact_count,
        "exact_verdict_agreement_rate": (
            exact_count / verdict_evaluable_count if verdict_evaluable_count else 0.0
        ),
        "false_disclosed_count": false_disclosed_count,
        "all_rows_false_disclosed_count": all_rows_false_disclosed_count,
        "unsupported_evidence_reference_count": unsupported_count,
        "wrong_source_page_count": wrong_page_count,
        "all_rows_wrong_source_page_count": all_rows_wrong_page_count,
        "manual_evidence_page_disagreement_count": manual_page_disagreement_count,
        "adjudication_pending_count": sum(
            case.manual.requirement_id in adjudication_pending_ids for case in cases
        ),
        "schema_failure_count": schema_failure_count,
        "model_failure_count": model_failure_count,
        "rules_ai_disagreement_count": disagreement_count,
        "manual_verdict_by_review_priority": dict(sorted(cross_distribution.items())),
    }
    return AIEvaluationResult(rows=rows, summary=summary)


def _read_dataset_identity(workbook) -> tuple[str, str]:
    values: dict[str, str] = {}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(min_row=1, max_row=30, values_only=True):
            if len(row) < 2:
                continue
            key = _text(row[0])
            if key in {"report_id", "run_id"}:
                values[key] = _text(row[1])
        if {"report_id", "run_id"}.issubset(values):
            break
    if not values.get("report_id") or not values.get("run_id"):
        raise ValueError("manual review workbook is missing report_id or run_id")
    return values["report_id"], values["run_id"]


def _find_review_sheet(workbook):
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
            start=1,
        ):
            values = [_text(value) for value in row]
            if "requirement_id" not in values:
                continue
            headers = {
                value: index for index, value in enumerate(values) if value
            }
            if set(MANUAL_REVIEW_FIELDS).issubset(headers):
                return worksheet, row_number, headers
    raise ValueError("manual review worksheet header was not found")


def _parse_pages(value: Any) -> list[int]:
    if value is None or value == "":
        return []
    if isinstance(value, (list, tuple)):
        return [int(page) for page in value]
    text = str(value).strip()
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        parsed = [part.strip() for part in text.split(",") if part.strip()]
    if isinstance(parsed, int):
        parsed = [parsed]
    return sorted({int(page) for page in parsed})


def _parse_json_pages(value: Any, *, requirement_id: str) -> list[int]:
    try:
        parsed = json.loads(_text(value))
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"invalid final PDF pages JSON: {requirement_id}"
        ) from exc
    if not isinstance(parsed, list) or any(
        not isinstance(page, int) or isinstance(page, bool) or page <= 0
        for page in parsed
    ):
        raise ValueError(
            f"final PDF pages must be a positive integer array: {requirement_id}"
        )
    if len(set(parsed)) != len(parsed):
        raise ValueError(f"duplicate final PDF pages: {requirement_id}")
    return sorted(parsed)


def _parse_json_strings(value: Any, *, requirement_id: str) -> list[str]:
    try:
        parsed = json.loads(_text(value))
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"invalid final missing items JSON: {requirement_id}"
        ) from exc
    if not isinstance(parsed, list) or any(
        not isinstance(item, str) or not item.strip() for item in parsed
    ):
        raise ValueError(
            f"final missing items must be a string array: {requirement_id}"
        )
    normalized = [item.strip() for item in parsed]
    if len(set(normalized)) != len(normalized):
        raise ValueError(f"duplicate final missing items: {requirement_id}")
    return normalized


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()

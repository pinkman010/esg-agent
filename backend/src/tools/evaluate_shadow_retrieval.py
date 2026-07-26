import argparse
import csv
import json
import re
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config.settings import PROJECT_ROOT, get_settings
from src.db.repositories import Repository
from src.db.session import SessionLocal
from src.standards.requirement_structure import canonical_requirement_id
from src.tools.embed_document_chunks import normalize_embedding_input
from src.tools.embedding_client import EmbeddingCallBlocked, EmbeddingClient
from src.tools.shadow_vector_retrieval import resolve_shadow_output


DEFAULT_REQUIREMENTS = (
    PROJECT_ROOT
    / "backend/data/manifests/gri_requirement_checklist_v3.json"
)
DEFAULT_BASELINE = (
    PROJECT_ROOT
    / "backend/data/review_inputs/envision_2024/baselines/"
    "current_577_review_regenerated.csv"
)
DEFAULT_MANUAL_REVIEW = (
    PROJECT_ROOT
    / "backend/data/review_inputs/envision_2024/manual/"
    "envision_2024_577_manual_review_second_review_Pro_20260719.xlsx"
)


def _ordered_unique(values: Iterable[int]) -> list[int]:
    seen: set[int] = set()
    result: list[int] = []
    for value in values:
        if value > 0 and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def parse_page_list(value: object) -> list[int]:
    if value is None or value == "":
        return []
    if isinstance(value, bool):
        return []
    if isinstance(value, int):
        return [value] if value > 0 else []
    if isinstance(value, float):
        return [int(value)] if value.is_integer() and value > 0 else []
    if isinstance(value, (list, tuple, set)):
        pages: list[int] = []
        for item in value:
            pages.extend(parse_page_list(item))
        return _ordered_unique(pages)

    text = str(value).strip()
    if not text or text == "[]":
        return []
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return _ordered_unique(
            int(match)
            for match in re.findall(r"\d+", text)
        )
    return parse_page_list(parsed)


def _case_metrics(
    case: dict[str, Any],
    *,
    k_values: tuple[int, ...],
) -> dict[str, Any]:
    gold_pages = set(parse_page_list(case.get("gold_pages")))
    vector_pages = _ordered_unique(
        parse_page_list(case.get("vector_pages"))
    )
    rule_pages = set(parse_page_list(case.get("rule_pages")))
    if not gold_pages:
        return {
            "first_hit_rank": None,
            "comparison_bucket": "not_evaluated",
            **{f"hit_at_{k}": None for k in k_values},
            **{f"recall_at_{k}": None for k in k_values},
        }

    first_hit_rank = next(
        (
            rank
            for rank, page in enumerate(vector_pages, start=1)
            if page in gold_pages
        ),
        None,
    )
    max_k = max(k_values)
    vector_hit = bool(set(vector_pages[:max_k]) & gold_pages)
    rule_hit = bool(rule_pages & gold_pages)
    if vector_hit and rule_hit:
        bucket = "both"
    elif vector_hit:
        bucket = "vector_only"
    elif rule_hit:
        bucket = "rule_only"
    else:
        bucket = "neither"

    metrics: dict[str, Any] = {
        "first_hit_rank": first_hit_rank,
        "comparison_bucket": bucket,
    }
    for k in k_values:
        matched = set(vector_pages[:k]) & gold_pages
        metrics[f"hit_at_{k}"] = bool(matched)
        metrics[f"recall_at_{k}"] = len(matched) / len(gold_pages)
    return metrics


def compute_retrieval_metrics(
    cases: list[dict[str, Any]],
    *,
    k_values: tuple[int, ...] = (1, 3, 5, 10),
) -> dict[str, int | float]:
    if not k_values or any(k < 1 for k in k_values):
        raise ValueError("k_values must contain positive integers")
    k_values = tuple(sorted(set(k_values)))
    evaluated = [
        (case, _case_metrics(case, k_values=k_values))
        for case in cases
        if parse_page_list(case.get("gold_pages"))
    ]
    denominator = len(evaluated)
    summary: dict[str, int | float] = {
        "case_count": len(cases),
        "evaluated_case_count": denominator,
        "no_gold_page_case_count": len(cases) - denominator,
    }
    for k in k_values:
        hit_total = sum(
            bool(metrics[f"hit_at_{k}"])
            for _, metrics in evaluated
        )
        recall_total = sum(
            float(metrics[f"recall_at_{k}"])
            for _, metrics in evaluated
        )
        summary[f"hit_at_{k}"] = (
            round(hit_total / denominator, 6)
            if denominator
            else 0.0
        )
        summary[f"recall_at_{k}"] = (
            round(recall_total / denominator, 6)
            if denominator
            else 0.0
        )
    reciprocal_rank_total = sum(
        1 / int(metrics["first_hit_rank"])
        for _, metrics in evaluated
        if metrics["first_hit_rank"] is not None
    )
    summary["mrr"] = (
        round(reciprocal_rank_total / denominator, 6)
        if denominator
        else 0.0
    )
    for bucket in ("vector_only", "rule_only", "both", "neither"):
        summary[f"{bucket}_hit_count"] = sum(
            metrics["comparison_bucket"] == bucket
            for _, metrics in evaluated
        )
    return summary


def load_requirement_queries(path: Path) -> list[dict[str, str]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    queries: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in payload.get("requirements", []):
        if item.get("evaluation_role") != "independent":
            continue
        requirement_id = canonical_requirement_id(
            str(item.get("requirement_id") or ""),
            str(item.get("canonical_disclosure_id") or "") or None,
        )
        if requirement_id in seen:
            raise ValueError(
                f"duplicate independent requirement: {requirement_id}"
            )
        query_text = str(
            item.get("effective_requirement_text")
            or item.get("source_requirement_text")
            or item.get("requirement_text")
            or ""
        ).strip()
        if not query_text:
            raise ValueError(
                f"empty requirement query: {requirement_id}"
            )
        seen.add(requirement_id)
        queries.append(
            {
                "requirement_id": requirement_id,
                "query_text": query_text,
            }
        )
    return queries


def load_rule_pages(path: Path) -> dict[str, list[int]]:
    pages_by_requirement: dict[str, list[int]] = {}
    with path.open(encoding="utf-8-sig", newline="") as file:
        for row in csv.DictReader(file):
            requirement_id = str(row.get("requirement_id") or "").strip()
            if not requirement_id:
                continue
            pages = pages_by_requirement.setdefault(requirement_id, [])
            pages.extend(
                parse_page_list(row.get("candidate_pdf_pages"))
            )
            pages.extend(parse_page_list(row.get("source_pdf_page")))
    return {
        requirement_id: _ordered_unique(pages)
        for requirement_id, pages in pages_by_requirement.items()
    }


def load_manual_gold(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        worksheet = workbook["人工复核577"]
        header_row = None
        headers: dict[str, int] = {}
        for row_number in range(1, min(20, worksheet.max_row) + 1):
            values = [
                worksheet.cell(row_number, column).value
                for column in range(1, worksheet.max_column + 1)
            ]
            if "requirement_id" in values and "correct_pdf_pages" in values:
                header_row = row_number
                headers = {
                    str(value): column
                    for column, value in enumerate(values, start=1)
                    if value
                }
                break
        if header_row is None:
            raise ValueError("manual review header row not found")

        gold: dict[str, dict[str, Any]] = {}
        for row_number in range(header_row + 1, worksheet.max_row + 1):
            requirement_id = str(
                worksheet.cell(
                    row_number,
                    headers["requirement_id"],
                ).value
                or ""
            ).strip()
            if not requirement_id:
                continue
            if requirement_id in gold:
                raise ValueError(
                    f"duplicate manual requirement: {requirement_id}"
                )
            gold[requirement_id] = {
                "gold_pages": parse_page_list(
                    worksheet.cell(
                        row_number,
                        headers["correct_pdf_pages"],
                    ).value
                ),
                "manual_suggested_verdict": str(
                    worksheet.cell(
                        row_number,
                        headers["suggested_verdict"],
                    ).value
                    or ""
                ).strip(),
                "manual_applicability": str(
                    worksheet.cell(
                        row_number,
                        headers["manual_applicability"],
                    ).value
                    or ""
                ).strip(),
                "standard_verified": str(
                    worksheet.cell(
                        row_number,
                        headers["standard_verified"],
                    ).value
                    or ""
                ).strip(),
                "review_complete": str(
                    worksheet.cell(
                        row_number,
                        headers["review_complete"],
                    ).value
                    or ""
                ).strip(),
            }
        return gold
    finally:
        workbook.close()


def _write_outputs(
    cases: list[dict[str, Any]],
    summary: dict[str, Any],
    *,
    output_prefix: str,
    k_values: tuple[int, ...],
) -> tuple[Path, Path]:
    cases_path = resolve_shadow_output(f"{output_prefix}_cases.csv")
    summary_path = resolve_shadow_output(f"{output_prefix}_summary.json")
    cases_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "report_id",
        "requirement_id",
        "query_text",
        "gold_pages",
        "rule_pages",
        "vector_pages",
        "vector_source_pages",
        "vector_chunk_ids",
        "vector_scores",
        "vector_texts",
        "manual_suggested_verdict",
        "manual_applicability",
        "standard_verified",
        "review_complete",
        "first_hit_rank",
        *[f"hit_at_{k}" for k in k_values],
        *[f"recall_at_{k}" for k in k_values],
        "comparison_bucket",
    ]
    with cases_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        for case in cases:
            annotation = _case_metrics(case, k_values=k_values)
            row = {**case, **annotation}
            writer.writerow(
                {
                    field: (
                        json.dumps(
                            row.get(field),
                            ensure_ascii=False,
                        )
                        if isinstance(row.get(field), (list, dict))
                        else row.get(field, "")
                    )
                    for field in fields
                }
            )
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return cases_path, summary_path


def evaluate_shadow_retrieval(
    *,
    report_id: str,
    requirements_path: Path,
    baseline_path: Path,
    manual_review_path: Path,
    top_k: int,
    output_prefix: str,
) -> dict[str, Any]:
    if top_k < 10:
        raise ValueError("top_k must be at least 10 for fixed metrics")
    settings = get_settings()
    if not settings.embedding_enabled:
        raise EmbeddingCallBlocked(
            "external embedding call requires EMBEDDING_ENABLED=true"
        )

    requirements = load_requirement_queries(requirements_path)
    rule_pages = load_rule_pages(baseline_path)
    manual_gold = load_manual_gold(manual_review_path)
    client = EmbeddingClient(
        model=settings.embedding_model,
        api_key=settings.embedding_api_key,
        base_url=settings.embedding_api_base,
        expected_dim=settings.embedding_dim,
        timeout_seconds=settings.embedding_timeout_seconds,
        max_retries=settings.embedding_max_retries,
        retry_delay_seconds=settings.embedding_retry_delay_seconds,
    )

    cases: list[dict[str, Any]] = []
    with SessionLocal() as session:
        repository = Repository(session)
        for start in range(
            0,
            len(requirements),
            settings.embedding_batch_size,
        ):
            batch = requirements[
                start : start + settings.embedding_batch_size
            ]
            query_texts = [
                normalize_embedding_input(
                    item["query_text"],
                    max_chars=settings.embedding_max_input_chars,
                )
                for item in batch
            ]
            embedding_result = client.embed_texts(
                query_texts,
                embedding_enabled=settings.embedding_enabled,
            )
            for item, query_text, query_embedding in zip(
                batch,
                query_texts,
                embedding_result.embeddings,
                strict=True,
            ):
                hits = repository.search_chunk_embeddings(
                    provider=settings.embedding_provider,
                    model=settings.embedding_model,
                    query_embedding=query_embedding,
                    report_id=report_id,
                    limit=top_k,
                )
                requirement_id = item["requirement_id"]
                manual = manual_gold.get(requirement_id, {})
                cases.append(
                    {
                        "report_id": report_id,
                        "requirement_id": requirement_id,
                        "query_text": query_text,
                        "gold_pages": manual.get("gold_pages", []),
                        "rule_pages": rule_pages.get(
                            requirement_id,
                            [],
                        ),
                        "vector_pages": _ordered_unique(
                            hit.source_page for hit in hits
                        ),
                        "vector_source_pages": [
                            hit.source_page for hit in hits
                        ],
                        "vector_chunk_ids": [
                            hit.chunk_id for hit in hits
                        ],
                        "vector_scores": [
                            round(hit.score, 8) for hit in hits
                        ],
                        "vector_texts": [hit.text for hit in hits],
                        "manual_suggested_verdict": manual.get(
                            "manual_suggested_verdict",
                            "",
                        ),
                        "manual_applicability": manual.get(
                            "manual_applicability",
                            "",
                        ),
                        "standard_verified": manual.get(
                            "standard_verified",
                            "",
                        ),
                        "review_complete": manual.get(
                            "review_complete",
                            "",
                        ),
                    }
                )

    k_values = (1, 3, 5, 10)
    summary: dict[str, Any] = {
        "report_id": report_id,
        "provider": settings.embedding_provider,
        "model": settings.embedding_model,
        "top_k": top_k,
        **compute_retrieval_metrics(cases, k_values=k_values),
    }
    cases_path, summary_path = _write_outputs(
        cases,
        summary,
        output_prefix=output_prefix,
        k_values=k_values,
    )
    return {
        **summary,
        "cases_output": str(cases_path),
        "summary_output": str(summary_path),
    }


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--report-id", required=True)
    parser.add_argument(
        "--requirements",
        type=Path,
        default=DEFAULT_REQUIREMENTS,
    )
    parser.add_argument(
        "--baseline",
        type=Path,
        default=DEFAULT_BASELINE,
    )
    parser.add_argument(
        "--manual-review-workbook",
        type=Path,
        default=DEFAULT_MANUAL_REVIEW,
    )
    parser.add_argument("--top-k", type=int, default=10)
    parser.add_argument(
        "--output-prefix",
        default="tmp/embedding/envision_shadow_retrieval",
    )
    args = parser.parse_args(argv)
    result = evaluate_shadow_retrieval(
        report_id=args.report_id,
        requirements_path=args.requirements,
        baseline_path=args.baseline,
        manual_review_path=args.manual_review_workbook,
        top_k=args.top_k,
        output_prefix=args.output_prefix,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

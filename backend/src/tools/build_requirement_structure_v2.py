from __future__ import annotations

import argparse
import csv
import json
import re
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.standards.requirement_structure import RequirementStructureDecision
from src.standards.requirement_structure import (
    EvaluationRole,
    canonical_requirement_id,
    compile_requirement_structure,
)


_SUPPORTED_ISSUE_CODES = {
    "parent_container_as_leaf",
    "missing_parent_context",
    "merge_split_error",
    "compound_requirement_adjudicated",
}
_PARENT_ID_PATTERN = re.compile(
    r"依赖上级\s*(GRI\s+\d+(?:-[A-Za-z0-9]+)+)\s*的",
    re.IGNORECASE,
)
EXPECTED_REVIEW_SHA256 = (
    "f1eeb37444de1eeda86b8ae0813dbfd6e88c94719781b98a8de659d9fbd7ddea"
)
EXPECTED_STRUCTURE_COUNTS = {
    "standard_unit_count": 577,
    "verified_count": 225,
    "context_only_count": 78,
    "normalized_count": 268,
    "method_pending_count": 6,
    "independent_assessment_count": 493,
}
EXPECTED_STRUCTURE_COUNTS_V3 = {
    "standard_unit_count": 577,
    "verified_count": 225,
    "context_only_count": 78,
    "normalized_count": 274,
    "method_pending_count": 0,
    "independent_assessment_count": 499,
    "compound_adjudicated_count": 6,
}


def read_structure_decisions(
    workbook_path: str | Path,
) -> list[RequirementStructureDecision]:
    path = Path(workbook_path)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if "人工复核577" not in workbook.sheetnames:
            raise ValueError("review workbook is missing sheet: 人工复核577")
        sheet = workbook["人工复核577"]
        rows = sheet.iter_rows(values_only=True)
        headers: list[str] | None = None
        decisions: list[RequirementStructureDecision] = []
        for values in rows:
            normalized = [str(value or "").strip() for value in values]
            if headers is None:
                if "requirement_id" not in normalized:
                    continue
                headers = normalized
                _require_headers(headers)
                continue
            row = dict(zip(headers, values, strict=False))
            requirement_id = _cell_text(row.get("requirement_id"))
            if not requirement_id:
                continue
            if _cell_text(row.get("standard_verified")).lower() != "no":
                continue
            review_note = _cell_text(row.get("review_note"))
            second_review_note = _cell_text(row.get("second_review_note"))
            source_note = "\n\n".join(
                note for note in (review_note, second_review_note) if note
            )
            issue_code = _classify_issue(
                explicit_code=_cell_text(row.get("primary_issue_type")),
                source_note=source_note,
                requirement_id=requirement_id,
            )
            parent_requirement_id = None
            if issue_code == "missing_parent_context":
                parent_requirement_id = _extract_parent_id(
                    source_note=source_note,
                    requirement_id=requirement_id,
                )
            decisions.append(
                RequirementStructureDecision(
                    requirement_id=requirement_id,
                    issue_code=issue_code,
                    parent_requirement_id=parent_requirement_id,
                    source_note=source_note,
                )
            )
        if headers is None:
            raise ValueError("review workbook header row was not found")
        return decisions
    finally:
        workbook.close()


def build_requirement_structure_assets(
    *,
    review_workbook: str | Path,
    source_checklist: str | Path,
    output_structure: str | Path,
    output_checklist: str | Path,
    structure_adjudication_csv: str | Path | None = None,
    manifest_version: str = "gri-requirement-checklist-v2",
    expected_review_sha256: str | None = EXPECTED_REVIEW_SHA256,
    expected_counts: dict[str, int] | None = EXPECTED_STRUCTURE_COUNTS,
) -> dict[str, Any]:
    review_path = Path(review_workbook)
    source_path = Path(source_checklist)
    review_hash = sha256(review_path.read_bytes()).hexdigest()
    if expected_review_sha256 and review_hash != expected_review_sha256:
        raise ValueError("review workbook SHA256 mismatch")

    source_bytes = source_path.read_bytes()
    source_hash = sha256(source_bytes).hexdigest()
    try:
        source_document = json.loads(source_bytes.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("invalid source checklist JSON") from exc
    if not isinstance(source_document, dict) or not isinstance(
        source_document.get("requirements"), list
    ):
        raise ValueError("source checklist must contain a requirements list")
    all_source_items: list[dict[str, Any]] = source_document["requirements"]
    source_items = [
        item for item in all_source_items if _is_current_scope_requirement(item)
    ]
    source_decisions = read_structure_decisions(review_path)
    _verify_parent_mappings(source_items=source_items, decisions=source_decisions)
    decisions = source_decisions
    adjudication_hash: str | None = None
    adjudication_rows: list[dict[str, str]] = []
    if structure_adjudication_csv is not None:
        adjudication_path = Path(structure_adjudication_csv)
        adjudication_hash = sha256(adjudication_path.read_bytes()).hexdigest()
        adjudication_rows = _read_structure_adjudications(adjudication_path)
        decisions = _apply_structure_adjudications(
            source_decisions=source_decisions,
            adjudication_rows=adjudication_rows,
        )
    compiled_items = compile_requirement_structure(
        items=source_items,
        decisions=decisions,
    )
    counts = _structure_counts(compiled_items=compiled_items, decisions=decisions)
    if expected_counts:
        mismatches = {
            key: {"expected": value, "actual": counts.get(key)}
            for key, value in expected_counts.items()
            if counts.get(key) != value
        }
        if mismatches:
            raise ValueError(
                f"structure count mismatch: {json.dumps(mismatches, sort_keys=True)}"
            )

    structure_manifest_version = _structure_manifest_version(manifest_version)
    structure_metadata = {
        "manifest_version": structure_manifest_version,
        "source_review_sha256": review_hash,
        "source_checklist_sha256": source_hash,
        "source_requirement_count": len(all_source_items),
        "excluded_source_requirement_count": len(all_source_items) - len(source_items),
        **counts,
    }
    if adjudication_hash:
        structure_metadata["source_adjudication_sha256"] = adjudication_hash
    structure_document = {
        "metadata": structure_metadata,
        "decisions": [decision.model_dump(mode="json") for decision in decisions],
        "provenance": {
            "source_review_decisions": [
                decision.model_dump(mode="json") for decision in source_decisions
            ],
            "structure_adjudications": adjudication_rows,
        },
    }
    compiled_document = {
        **source_document,
        "metadata": {
            **dict(source_document.get("metadata") or {}),
            "manifest_version": manifest_version,
            "source_review_sha256": review_hash,
            "source_checklist_sha256": source_hash,
            "structure_manifest_version": structure_metadata["manifest_version"],
            **counts,
        },
        "requirements": compiled_items,
    }
    if adjudication_hash:
        compiled_document["metadata"]["source_adjudication_sha256"] = adjudication_hash
    _write_json_pair_atomically(
        output_structure=Path(output_structure),
        structure_document=structure_document,
        output_checklist=Path(output_checklist),
        checklist_document=compiled_document,
    )
    return structure_metadata


def _verify_parent_mappings(
    *,
    source_items: list[dict[str, Any]],
    decisions: list[RequirementStructureDecision],
) -> None:
    source_by_id: dict[str, dict[str, Any]] = {}
    for item in source_items:
        canonical_id = canonical_requirement_id(
            str(item.get("requirement_id") or ""),
            _cell_text(item.get("canonical_disclosure_id")) or None,
        )
        if canonical_id in source_by_id:
            raise ValueError(f"duplicate requirement_id: {canonical_id}")
        source_by_id[canonical_id] = item
    for decision in decisions:
        if decision.issue_code != "missing_parent_context":
            continue
        item = source_by_id.get(decision.requirement_id)
        if item is None:
            raise ValueError(
                f"structure decision has no checklist item: {decision.requirement_id}"
            )
        raw_parent_id = _cell_text(item.get("parent_requirement_id"))
        if not raw_parent_id:
            raise ValueError(f"source parent is missing: {decision.requirement_id}")
        source_parent_id = canonical_requirement_id(
            raw_parent_id,
            _cell_text(item.get("canonical_disclosure_id")) or None,
        )
        if source_parent_id == decision.parent_requirement_id:
            continue
        disclosure_id = _cell_text(item.get("canonical_disclosure_id"))
        disclosure_parent_id = f"GRI {disclosure_id}" if disclosure_id else ""
        direct_parent = source_by_id.get(decision.parent_requirement_id)
        lexical_parent_id = decision.requirement_id.rsplit("-", 1)[0]
        flat_pointer_is_consistent = (
            source_parent_id == disclosure_parent_id
            and lexical_parent_id == decision.parent_requirement_id
            and direct_parent is not None
            and _cell_text(direct_parent.get("canonical_disclosure_id")) == disclosure_id
        )
        if not flat_pointer_is_consistent:
            raise ValueError(
                "parent mapping mismatch: "
                f"{decision.requirement_id} review={decision.parent_requirement_id} "
                f"source={source_parent_id}"
            )


def _is_current_scope_requirement(item: dict[str, Any]) -> bool:
    return (
        item.get("assessment_mode") == "current_gap"
        and item.get("requirement_type") == "requirement"
        and item.get("is_mandatory") is True
        and item.get("scoring_role") == "hard_score"
    )


def _structure_counts(
    *,
    compiled_items: list[dict[str, Any]],
    decisions: list[RequirementStructureDecision],
) -> dict[str, int]:
    issue_counts = {
        issue_code: sum(item.issue_code == issue_code for item in decisions)
        for issue_code in _SUPPORTED_ISSUE_CODES
    }
    independent_count = sum(
        item.get("evaluation_role") == EvaluationRole.INDEPENDENT.value
        for item in compiled_items
    )
    return {
        "standard_unit_count": len(compiled_items),
        "verified_count": len(compiled_items) - len(decisions),
        "context_only_count": issue_counts["parent_container_as_leaf"],
        "normalized_count": (
            issue_counts["missing_parent_context"]
            + issue_counts["compound_requirement_adjudicated"]
        ),
        "method_pending_count": issue_counts["merge_split_error"],
        "compound_adjudicated_count": issue_counts[
            "compound_requirement_adjudicated"
        ],
        "independent_assessment_count": independent_count,
    }


def _read_structure_adjudications(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required_headers = {
            "requirement_id",
            "final_evaluation_role",
            "component_requirement_ids",
            "completeness_policy",
            "adjudication_rationale",
            "adjudicator_role",
            "adjudication_version",
        }
        missing_headers = sorted(required_headers.difference(reader.fieldnames or []))
        if missing_headers:
            raise ValueError(
                "structure adjudication CSV missing columns: "
                + ", ".join(missing_headers)
            )
        rows = [
            {key: _cell_text(value) for key, value in row.items()}
            for row in reader
            if _cell_text(row.get("requirement_id"))
        ]
    return rows


def _apply_structure_adjudications(
    *,
    source_decisions: list[RequirementStructureDecision],
    adjudication_rows: list[dict[str, str]],
) -> list[RequirementStructureDecision]:
    pending_by_id = {
        decision.requirement_id: decision
        for decision in source_decisions
        if decision.issue_code == "merge_split_error"
    }
    row_by_id: dict[str, dict[str, str]] = {}
    for row in adjudication_rows:
        requirement_id = row["requirement_id"]
        if requirement_id in row_by_id:
            raise ValueError(f"duplicate structure adjudication: {requirement_id}")
        if requirement_id not in pending_by_id:
            raise ValueError(
                "structure adjudication is not a merge_split_error: "
                f"{requirement_id}"
            )
        if row["final_evaluation_role"] != EvaluationRole.INDEPENDENT.value:
            raise ValueError(
                "compound structure adjudication must be independent: "
                f"{requirement_id}"
            )
        row_by_id[requirement_id] = row
    missing_ids = sorted(set(pending_by_id).difference(row_by_id))
    if missing_ids:
        raise ValueError(
            "structure adjudication is missing merge_split_error IDs: "
            + ", ".join(missing_ids)
        )

    resolved: list[RequirementStructureDecision] = []
    for decision in source_decisions:
        row = row_by_id.get(decision.requirement_id)
        if row is None:
            resolved.append(decision)
            continue
        try:
            component_ids = json.loads(row["component_requirement_ids"])
        except json.JSONDecodeError as exc:
            raise ValueError(
                "invalid component_requirement_ids JSON: "
                f"{decision.requirement_id}"
            ) from exc
        if not isinstance(component_ids, list) or not all(
            isinstance(item, str) and item.strip() for item in component_ids
        ):
            raise ValueError(
                "component_requirement_ids must be a JSON string array: "
                f"{decision.requirement_id}"
            )
        resolved.append(
            RequirementStructureDecision(
                requirement_id=decision.requirement_id,
                issue_code="compound_requirement_adjudicated",
                parent_requirement_id=decision.parent_requirement_id,
                source_note=row["adjudication_rationale"],
                evaluation_role=EvaluationRole.INDEPENDENT,
                component_requirement_ids=[item.strip() for item in component_ids],
                adjudication_version=row["adjudication_version"],
            )
        )
    return resolved


def _structure_manifest_version(checklist_manifest_version: str) -> str:
    prefix = "gri-requirement-checklist-"
    if not checklist_manifest_version.startswith(prefix):
        raise ValueError(
            f"unsupported checklist manifest version: {checklist_manifest_version}"
        )
    return checklist_manifest_version.replace(
        "gri-requirement-checklist-",
        "gri-requirement-structure-",
        1,
    )


def _expected_counts_for_manifest(manifest_version: str) -> dict[str, int]:
    if manifest_version == "gri-requirement-checklist-v2":
        return dict(EXPECTED_STRUCTURE_COUNTS)
    if manifest_version == "gri-requirement-checklist-v3":
        return dict(EXPECTED_STRUCTURE_COUNTS_V3)
    raise ValueError(f"unsupported checklist manifest version: {manifest_version}")


def _write_json_pair_atomically(
    *,
    output_structure: Path,
    structure_document: dict[str, Any],
    output_checklist: Path,
    checklist_document: dict[str, Any],
) -> None:
    output_structure.parent.mkdir(parents=True, exist_ok=True)
    output_checklist.parent.mkdir(parents=True, exist_ok=True)
    structure_temp = output_structure.with_suffix(f"{output_structure.suffix}.tmp")
    checklist_temp = output_checklist.with_suffix(f"{output_checklist.suffix}.tmp")
    structure_temp.write_text(
        json.dumps(structure_document, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    checklist_temp.write_text(
        json.dumps(checklist_document, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    structure_temp.replace(output_structure)
    checklist_temp.replace(output_checklist)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compile versioned reviewed GRI requirement structure assets."
    )
    parser.add_argument("--review-workbook", required=True, type=Path)
    parser.add_argument("--source-checklist", required=True, type=Path)
    parser.add_argument("--output-structure", required=True, type=Path)
    parser.add_argument("--output-checklist", required=True, type=Path)
    parser.add_argument("--structure-adjudication-csv", type=Path)
    parser.add_argument(
        "--manifest-version",
        default="gri-requirement-checklist-v3",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    metadata = build_requirement_structure_assets(
        review_workbook=args.review_workbook,
        source_checklist=args.source_checklist,
        output_structure=args.output_structure,
        output_checklist=args.output_checklist,
        structure_adjudication_csv=args.structure_adjudication_csv,
        manifest_version=args.manifest_version,
        expected_counts=_expected_counts_for_manifest(args.manifest_version),
    )
    print(json.dumps(metadata, ensure_ascii=False, sort_keys=True))
    return 0


def _require_headers(headers: list[str]) -> None:
    required = {
        "requirement_id",
        "standard_verified",
        "review_note",
        "second_review_note",
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"review workbook missing columns: {', '.join(missing)}")


def _classify_issue(
    *,
    explicit_code: str,
    source_note: str,
    requirement_id: str,
) -> str:
    if explicit_code:
        if explicit_code not in _SUPPORTED_ISSUE_CODES:
            raise ValueError(
                f"unsupported structure issue code for {requirement_id}: {explicit_code}"
            )
        return explicit_code
    if "父级/容器" in source_note:
        return "parent_container_as_leaf"
    if "依赖上级" in source_note:
        return "missing_parent_context"
    if "合并" in source_note and ("拆分" in source_note or "并列" in source_note):
        return "merge_split_error"
    raise ValueError(f"cannot classify structure issue: {requirement_id}")


def _extract_parent_id(*, source_note: str, requirement_id: str) -> str:
    match = _PARENT_ID_PATTERN.search(source_note)
    if match is None:
        raise ValueError(f"cannot extract parent requirement_id: {requirement_id}")
    return " ".join(match.group(1).split())


def _cell_text(value: object) -> str:
    return str(value or "").strip()


if __name__ == "__main__":
    raise SystemExit(main())

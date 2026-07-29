from hashlib import sha256
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sqlalchemy import select

from src.config.settings import get_settings
from src.db.models import AssessmentRecord, AuditEventRecord, ExportVersionRecord
from src.db.repositories import Repository
from src.domain.ai_models import AIAssessmentSuggestion
from src.domain.enums import AISuggestionStatus, AssessmentVerdict, EvidenceSourceMethod, PageQualityFlag, ReportStatus, ReviewOperation, ReviewStatus, RunStatus
from src.domain.models import AnalysisRun, DisclosureAssessment, DisclosureTask, EvidenceItem, Report, ReviewDecision
from src.services.risk_service import calculate_and_store_risk
from src.services.review_service import ReviewService
from src.standards.gri import GRIAdapter

pytestmark = pytest.mark.anyio


async def download_export_file(api_client, export: dict, format_name: str):
    item = next(
        item
        for item in export["file_manifest"]
        if item["format"] == format_name
    )
    response = await api_client.get(
        f"/api/exports/{export['export_id']}/files/{item['file_id']}"
    )
    return item, response


def seed_export_data(session):
    repo = Repository(session)
    repo.create_report(Report(report_id="report-1", original_filename="report.pdf", stored_path="x", file_hash="hash-1"))
    repo.create_run(AnalysisRun(run_id="run-1", report_id="report-1", status=RunStatus.COMPLETED))
    assessment = DisclosureAssessment(
        assessment_id="assessment-1",
        run_id="run-1",
        report_id="report-1",
        standard_id="GRI",
        standard_version="2021",
        disclosure_id="GRI 302",
        requirement_id="GRI 302-1-a",
        verdict=AssessmentVerdict.DISCLOSED,
        rationale="The report index contains an omission note, but no substantive disclosure evidence was found.",
        missing_items=[
            "EVG&D source basis from audited financial/P&L statement or internally audited management accounts",
            "applicability of EVG&D source basis",
        ],
        evidence=[
            EvidenceItem(
                evidence_id="evidence-1",
                run_id="run-1",
                report_id="report-1",
                source_text="独立有限鉴证报告",
                source_page=77,
                source_pdf_page=77,
                source_report_page=76,
                source_file_hash="hash-1",
                source_method=EvidenceSourceMethod.PDFPLUMBER,
                quality_flags=[PageQualityFlag.SHORT_TEXT, PageQualityFlag.IMAGE_BODY_NOT_EXTRACTED],
                needs_ocr_or_vlm=True,
                requires_ocr=True,
                requires_vlm=False,
                ocr_or_vlm_reason="assurance_page_text_too_short",
                metadata={
                    "candidate_pdf_pages": [77],
                    "candidate_report_pages": [76],
                },
            )
        ],
        review_status=ReviewStatus.NOT_REQUIRED,
    )
    repo.save_assessment(assessment)
    repo.save_evidence_item("assessment-1", assessment.evidence[0])
    repo.save_disclosure_task(
        DisclosureTask(
            task_id="task-1",
            run_id="run-1",
            report_id="report-1",
            standard_id="GRI",
            standard_version="2021",
            disclosure_id="GRI 302",
            requirement_id="GRI 302-1-a",
            requirement_text="披露组织内部能源消耗量。",
            source_requirement_text="组织内部能源消耗量",
            context_requirement_ids=["GRI 302-1"],
            structure_status="verified",
        )
    )
    repo.append_ai_suggestion(
        AIAssessmentSuggestion(
            suggestion_id="ai-suggestion-1",
            assessment_id="assessment-1",
            run_id="run-1",
            status=AISuggestionStatus.SUCCEEDED,
            provider="deepseek",
            model="deepseek-v4-flash",
            prompt_version="deepseek-gri-assist-v1",
            input_hash="a" * 64,
            suggested_verdict=AssessmentVerdict.UNKNOWN,
            rationale_zh="输入证据不足以支持完整披露。",
            missing_items_zh=["实质披露内容"],
            evidence_ids=["evidence-1"],
            evidence_pdf_pages=[77],
            confidence=0.6,
        )
    )
    calculate_and_store_risk(repo, assessment, trigger_event="analysis_completed")
    repo.save_review_decision(ReviewDecision(decision_id="decision-1", run_id="run-1", assessment_id="assessment-1", review_status=ReviewStatus.APPROVED, reviewer_note="Checked."))


async def test_export_api_returns_json_and_csv(api_client, api_session):
    seed_export_data(api_session)

    assessments_json = await api_client.get("/api/exports/runs/run-1/assessments.json")
    assessments_csv = await api_client.get("/api/exports/runs/run-1/assessments.csv")
    review_json = await api_client.get("/api/exports/runs/run-1/review.json")
    review_csv = await api_client.get("/api/exports/runs/run-1/review.csv")

    assert assessments_json.status_code == 200
    assert assessments_json.json()[0]["assessment_id"] == "assessment-1"
    assert assessments_json.json()[0]["source_pdf_page"] == 77
    assert assessments_json.json()[0]["source_report_page"] == 76
    assert assessments_json.json()[0]["page_label"] == "PDF 第 77 页 / 报告页 76"
    assert assessments_json.json()[0]["needs_ocr_or_vlm"] is True
    assert assessments_json.json()[0]["requires_ocr"] is True
    assert assessments_json.json()[0]["requires_vlm"] is False
    assert assessments_json.json()[0]["evidence_preview"] == "独立有限鉴证报告"
    assert assessments_json.json()[0]["candidate_pdf_pages"] == [77]
    assert assessments_json.json()[0]["candidate_report_pages"] == [76]
    assert assessments_json.json()[0]["rationale"].startswith("The report index contains")
    assert assessments_json.json()[0]["rationale_zh"] == "报告 GRI 内容索引包含从略说明，但未找到实质性披露证据。"
    assert assessments_json.json()[0]["missing_items"][0].startswith("EVG&D source basis")
    assert assessments_json.json()[0]["missing_items_zh"][1] == "EVG&D 数据来源依据的适用性说明"
    assert assessments_json.json()[0]["structure_status"] == "verified"
    assert assessments_json.json()[0]["source_requirement_text"] == "组织内部能源消耗量"
    assert assessments_json.json()[0]["effective_requirement_text"] == "披露组织内部能源消耗量。"
    assert assessments_json.json()[0]["ai_status"] == "succeeded"
    assert assessments_json.json()[0]["ai_suggested_verdict"] == "unknown"
    assert assessments_json.json()[0]["ai_evidence_pdf_pages"] == [77]
    assert assessments_json.json()[0]["ai_model"] == "deepseek-v4-flash"
    assert assessments_csv.status_code == 200
    assert "assessment_id" in assessments_csv.text
    assert "source_pdf_page" in assessments_csv.text
    assert "source_report_page" in assessments_csv.text
    assert "page_label" in assessments_csv.text
    assert "PDF 第 77 页 / 报告页 76" in assessments_csv.text
    assert "needs_ocr_or_vlm" in assessments_csv.text
    assert "requires_ocr" in assessments_csv.text
    assert "requires_vlm" in assessments_csv.text
    assert "evidence_preview" in assessments_csv.text
    assert "candidate_pdf_pages" in assessments_csv.text
    assert "candidate_report_pages" in assessments_csv.text
    assert "rationale_zh" in assessments_csv.text
    assert "missing_items_zh" in assessments_csv.text
    assert "ai_suggested_verdict" in assessments_csv.text
    assert review_json.json()[0]["decision_id"] == "decision-1"
    assert "decision_id" in review_csv.text
    event_types = api_session.scalars(
        select(AuditEventRecord.event_type)
        .where(AuditEventRecord.run_id == "run-1")
        .order_by(AuditEventRecord.audit_event_id)
    ).all()
    assert event_types == [
        "assessments_json_exported",
        "assessments_csv_exported",
        "review_json_exported",
        "review_csv_exported",
    ]


async def test_assessment_xlsx_places_ai_disclaimer_before_headers(
    api_client,
    api_session,
):
    seed_export_data(api_session)

    response = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )

    assert response.status_code == 200
    body = response.json()
    item = body["file_manifest"][0]
    assert set(item) == {
        "file_id",
        "filename",
        "format",
        "size",
        "sha256",
    }
    assert "path" not in response.text
    assert "relative_path" not in response.text

    item, download = await download_export_file(
        api_client,
        body,
        "assessment_xlsx",
    )
    assert download.status_code == 200
    assert download.headers["content-disposition"].startswith("attachment;")
    assert len(download.content) == item["size"]
    assert sha256(download.content).hexdigest() == item["sha256"]

    workbook = load_workbook(BytesIO(download.content), read_only=True)
    sheet = workbook["GRI核查"]
    assert "AI建议未经人工确认时不构成最终披露结论" in sheet["A1"].value
    headers = [cell.value for cell in sheet[2]]
    assert "structure_status" in headers
    assert "ai_suggested_verdict" in headers
    workbook.close()
    event = api_session.scalar(
        select(AuditEventRecord)
        .where(AuditEventRecord.event_type == "export_file_downloaded")
        .order_by(AuditEventRecord.audit_event_id.desc())
        .limit(1)
    )
    assert event.run_id == "run-1"
    assert event.event_payload == {
        "export_id": body["export_id"],
        "file_id": item["file_id"],
        "format": "assessment_xlsx",
        "size": item["size"],
        "sha256": item["sha256"],
    }


async def test_export_download_returns_stable_safe_errors(
    api_client,
    api_session,
):
    seed_export_data(api_session)
    settings = get_settings()

    missing_export = await api_client.get(
        "/api/exports/export-missing/files/file-missing"
    )
    assert missing_export.status_code == 404
    assert missing_export.json()["detail"] == "export_not_found"

    generated = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )
    body = generated.json()
    item = body["file_manifest"][0]
    record = api_session.get(ExportVersionRecord, body["export_id"])
    original_manifest = [dict(entry) for entry in record.file_manifest]

    unknown_file = await api_client.get(
        f"/api/exports/{body['export_id']}/files/file-missing"
    )
    assert unknown_file.status_code == 404
    assert unknown_file.json()["detail"] == "export_file_not_found"

    outside_path = settings.derived_dir / "outside.xlsx"
    outside_path.parent.mkdir(parents=True, exist_ok=True)
    outside_path.write_bytes(b"outside")
    outside_manifest = dict(original_manifest[0])
    outside_manifest["relative_path"] = "outside.xlsx"
    record.file_manifest = [outside_manifest]
    api_session.commit()
    outside = await api_client.get(
        f"/api/exports/{body['export_id']}/files/{item['file_id']}"
    )
    assert outside.status_code == 404
    assert outside.json()["detail"] == "export_file_not_found"
    assert str(settings.derived_dir) not in outside.text

    missing_manifest = dict(original_manifest[0])
    missing_manifest["relative_path"] = (
        f"exports/report-1/{body['export_id']}/missing.xlsx"
    )
    record.file_manifest = [missing_manifest]
    api_session.commit()
    missing = await api_client.get(
        f"/api/exports/{body['export_id']}/files/{item['file_id']}"
    )
    assert missing.status_code == 404
    assert missing.json()["detail"] == "export_file_not_found"
    assert str(settings.derived_dir) not in missing.text

    record.file_manifest = original_manifest
    api_session.commit()
    original_path = settings.derived_dir / original_manifest[0]["relative_path"]
    original_path.write_bytes(original_path.read_bytes() + b"tampered")
    mismatch = await api_client.get(
        f"/api/exports/{body['export_id']}/files/{item['file_id']}"
    )
    assert mismatch.status_code == 409
    assert mismatch.json()["detail"] == "export_file_integrity_mismatch"
    assert str(settings.derived_dir) not in mismatch.text


async def test_legacy_export_manifest_is_safely_projected_and_downloadable(
    api_client,
    api_session,
):
    seed_export_data(api_session)
    settings = get_settings()
    generated = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )
    export_id = generated.json()["export_id"]
    record = api_session.get(ExportVersionRecord, export_id)
    internal = record.file_manifest[0]
    legacy_path = settings.derived_dir / internal["relative_path"]
    record.file_manifest = [
        {
            "format": internal["format"],
            "path": legacy_path.as_posix(),
            "size": legacy_path.stat().st_size,
            "sha256": sha256(legacy_path.read_bytes()).hexdigest(),
        }
    ]
    api_session.commit()

    listed = await api_client.get("/api/reports/report-1/exports")
    listed_export = next(
        item for item in listed.json() if item["export_id"] == export_id
    )
    item = listed_export["file_manifest"][0]
    assert set(item) == {
        "file_id",
        "filename",
        "format",
        "size",
        "sha256",
    }
    assert "path" not in listed.text
    assert "relative_path" not in listed.text

    first = await api_client.get(
        f"/api/exports/{export_id}/files/{item['file_id']}"
    )
    second_list = await api_client.get("/api/reports/report-1/exports")
    second_export = next(
        entry for entry in second_list.json()
        if entry["export_id"] == export_id
    )
    assert second_export["file_manifest"][0]["file_id"] == item["file_id"]
    assert first.status_code == 200
    assert first.content == legacy_path.read_bytes()


async def test_versioned_export_generates_actions_xlsx(
    api_client,
    api_session,
):
    seed_export_data(api_session)
    created = await api_client.post(
        "/api/reports/report-1/actions",
        json={
            "assessment_id": "assessment-1",
            "title": '=HYPERLINK("https://example.invalid","打开")',
            "priority": "high",
            "owner_name": "+1+1",
            "due_date": "2026-08-31",
            "recommendation_text": "@SUM(1,1)",
            "created_by": "-1+1",
        },
    )

    response = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": ["actions_xlsx"], "created_by": "张三"},
    )

    assert created.status_code == 200
    assert response.status_code == 200
    body = response.json()
    item, download = await download_export_file(
        api_client,
        body,
        "actions_xlsx",
    )
    assert item["filename"] == "actions_xlsx.xlsx"
    assert download.status_code == 200
    workbook = load_workbook(BytesIO(download.content), read_only=True)
    sheet = workbook["整改任务"]
    assert sheet["A1"].value == (
        "本清单为整改任务跟踪材料；任务状态和截止日期"
        "不构成 GRI 认证或外部鉴证结论。"
    )
    assert [cell.value for cell in sheet[2]] == [
        "整改任务 ID",
        "Requirement ID",
        "任务标题",
        "优先级",
        "状态",
        "负责人",
        "截止日期",
        "建议内容",
        "完成说明",
        "创建人",
        "创建时间",
        "更新时间",
    ]
    row = [cell.value for cell in sheet[3]]
    assert row[:7] == [
        created.json()["action_id"],
        "GRI 302-1-a",
        '\'=HYPERLINK("https://example.invalid","打开")',
        "high",
        "open",
        "'+1+1",
        "2026-08-31",
    ]
    assert row[7:10] == [
        "'@SUM(1,1)",
        None,
        "'-1+1",
    ]
    for column in ("C", "F", "H", "J"):
        assert sheet[f"{column}3"].data_type == "s"
    workbook.close()


async def test_actions_xlsx_is_valid_when_report_has_no_actions(
    api_client,
    api_session,
):
    seed_export_data(api_session)

    response = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": ["actions_xlsx"], "created_by": "张三"},
    )
    _, download = await download_export_file(
        api_client,
        response.json(),
        "actions_xlsx",
    )

    assert response.status_code == 200
    assert download.status_code == 200
    workbook = load_workbook(BytesIO(download.content), read_only=True)
    sheet = workbook["整改任务"]
    assert sheet.max_row == 2
    assert sheet["A1"].value.startswith("本清单为整改任务跟踪材料")
    assert sheet["A2"].value == "整改任务 ID"
    workbook.close()


@pytest.mark.parametrize(
    "formats",
    [
        [],
        ["assessment_xlsx", "assessment_xlsx"],
    ],
)
async def test_versioned_export_rejects_empty_or_duplicate_formats(
    api_client,
    api_session,
    formats,
):
    seed_export_data(api_session)

    response = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": formats, "created_by": "张三"},
    )

    assert response.status_code == 422


async def test_versioned_draft_and_formal_exports(api_client, api_session):
    seed_export_data(api_session)

    draft = await api_client.post(
        "/api/reports/report-1/exports/draft",
        json={"formats": ["assessment_xlsx", "management_pdf", "print_html"], "created_by": "张三"},
    )
    blocked = await api_client.post(
        "/api/reports/report-1/exports/formal",
        json={"formats": ["actions_xlsx"], "created_by": "张三"},
    )
    ReviewService(Repository(api_session)).record(
        "assessment-1",
        operation_type=ReviewOperation.APPROVE,
        reviewer_name="张三",
        reason_code="system_result_confirmed",
    )
    formal = await api_client.post(
        "/api/reports/report-1/exports/formal",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )
    second_formal = await api_client.post(
        "/api/reports/report-1/exports/formal",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )
    listed = await api_client.get("/api/reports/report-1/exports")

    assert draft.status_code == 200
    assert draft.json()["is_draft"] is True
    assert draft.json()["review_scope"]["draft_label"] is True
    assert "高复核优先级未处理 1 条" in draft.json()["review_scope"]["review_scope_statement"]
    assert len(draft.json()["file_manifest"]) == 3
    assert blocked.status_code == 409
    assert blocked.json()["detail"] == {"code": "high_risk_review_incomplete", "remaining": 1}
    assert formal.status_code == 200
    assert formal.json()["version_number"] == 1
    assert formal.json()["is_draft"] is False
    assert second_formal.status_code == 200
    assert second_formal.json()["version_number"] == 2
    assert second_formal.json()["supersedes_export_id"] == formal.json()["export_id"]
    assert len(listed.json()) == 3
    first_formal = next(item for item in listed.json() if item["export_id"] == formal.json()["export_id"])
    assert first_formal["status"] == "superseded"
    assert Repository(api_session).get_report("report-1").status is ReportStatus.FORMALLY_EXPORTED


def seed_risk_v2_1_export_scope(session):
    repo = Repository(session)
    repo.create_report(
        Report(
            report_id="report-v2-1",
            original_filename="report.pdf",
            stored_path="x",
            file_hash="hash-v2-1",
        )
    )
    repo.create_run(
        AnalysisRun(
            run_id="run-v2-1",
            report_id="report-v2-1",
            status=RunStatus.COMPLETED,
            risk_rule_version="risk-v2.1",
            eligible_requirement_count=3,
            succeeded_requirement_count=3,
        )
    )
    cases = [
        (
            "assessment-high",
            AssessmentVerdict.DISCLOSED,
            "omission_note",
        ),
        (
            "assessment-medium",
            AssessmentVerdict.UNKNOWN,
            "index_statement",
        ),
        (
            "assessment-low",
            AssessmentVerdict.UNKNOWN,
            None,
        ),
    ]
    for index, (assessment_id, verdict, evidence_type) in enumerate(cases, start=1):
        evidence = []
        if evidence_type is not None:
            evidence = [
                EvidenceItem(
                    evidence_id=f"evidence-v2-1-{index}",
                    run_id="run-v2-1",
                    report_id="report-v2-1",
                    source_text="索引或从略说明",
                    source_page=72,
                    source_file_hash="hash-v2-1",
                    source_method=EvidenceSourceMethod.PDFPLUMBER,
                    metadata={"evidence_type": evidence_type},
                )
            ]
        assessment = DisclosureAssessment(
            assessment_id=assessment_id,
            run_id="run-v2-1",
            report_id="report-v2-1",
            standard_id="GRI",
            standard_version="2021",
            disclosure_id=f"GRI 2-{index}",
            requirement_id=f"GRI 2-{index}-a",
            verdict=verdict,
            rationale="待人工核实",
            evidence=evidence,
            review_status=ReviewStatus.NEEDS_MANUAL_REVIEW,
        )
        repo.save_assessment(assessment)
        for item in evidence:
            repo.save_evidence_item(assessment_id, item)
        calculate_and_store_risk(
            repo,
            assessment,
            trigger_event="analysis_completed",
            risk_rule_version="risk-v2.1",
        )


async def test_risk_v2_1_formal_export_blocks_only_unresolved_high_and_discloses_scope(
    api_client,
    api_session,
):
    seed_risk_v2_1_export_scope(api_session)

    blocked = await api_client.post(
        "/api/reports/report-v2-1/exports/formal",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )
    ReviewService(Repository(api_session)).record(
        "assessment-high",
        operation_type=ReviewOperation.APPROVE,
        reviewer_name="张三",
        reason_code="system_result_confirmed",
    )
    formal = await api_client.post(
        "/api/reports/report-v2-1/exports/formal",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )

    assert blocked.status_code == 409
    assert blocked.json()["detail"]["remaining"] == 1
    assert formal.status_code == 200
    scope = formal.json()["review_scope"]
    assert scope["high_priority_total"] == 1
    assert scope["high_priority_reviewed"] == 1
    assert scope["high_priority_unresolved"] == 0
    assert scope["medium_priority_total"] == 1
    assert scope["medium_priority_unresolved"] == 1
    assert scope["applicability_undetermined_total"] == 2
    assert scope["eligible_requirement_total"] == 3
    assert scope["human_reviewed_total"] == 1
    assert "不代表全部 3 条均已人工确认" in scope["review_scope_statement"]


async def test_risk_v2_1_formal_export_blocks_incomplete_analysis(api_client, api_session):
    repo = Repository(api_session)
    repo.create_report(
        Report(
            report_id="report-incomplete",
            original_filename="report.pdf",
            stored_path="x",
            file_hash="hash-incomplete",
        )
    )
    repo.create_run(
        AnalysisRun(
            run_id="run-incomplete",
            report_id="report-incomplete",
            status=RunStatus.PARTIALLY_COMPLETED,
            risk_rule_version="risk-v2.1",
            eligible_requirement_count=3,
            succeeded_requirement_count=2,
            failed_requirement_count=1,
        )
    )
    for index in range(1, 3):
        assessment = DisclosureAssessment(
            assessment_id=f"assessment-incomplete-{index}",
            run_id="run-incomplete",
            report_id="report-incomplete",
            standard_id="GRI",
            standard_version="2021",
            disclosure_id=f"GRI 2-{index}",
            requirement_id=f"GRI 2-{index}-a",
            verdict=AssessmentVerdict.UNKNOWN,
            rationale="No valid evidence was found.",
            review_status=ReviewStatus.NEEDS_MANUAL_REVIEW,
        )
        repo.save_assessment(assessment)
        calculate_and_store_risk(
            repo,
            assessment,
            trigger_event="analysis_completed",
            risk_rule_version="risk-v2.1",
        )

    draft = await api_client.post(
        "/api/reports/report-incomplete/exports/draft",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )
    formal = await api_client.post(
        "/api/reports/report-incomplete/exports/formal",
        json={"formats": ["assessment_xlsx"], "created_by": "张三"},
    )

    assert draft.status_code == 200
    assert draft.json()["review_scope"]["analysis_incomplete_total"] == 1
    assert "分析失败或未生成结果 1 条" in draft.json()["review_scope"]["review_scope_statement"]
    assert formal.status_code == 409
    assert formal.json()["detail"] == {
        "code": "analysis_incomplete",
        "remaining": 1,
    }


def seed_complete_v3_export_scope(session):
    repo = Repository(session)
    repo.create_report(
        Report(
            report_id="report-v3-export",
            original_filename="envision.pdf",
            stored_path="x",
            file_hash="hash-v3-export",
            status=ReportStatus.ANALYSIS_COMPLETED,
        )
    )
    repo.create_run(
        AnalysisRun(
            run_id="run-v3-export",
            report_id="report-v3-export",
            status=RunStatus.COMPLETED,
            risk_rule_version="risk-v2.1",
            eligible_requirement_count=499,
            succeeded_requirement_count=499,
        )
    )
    backend_root = Path(__file__).resolve().parents[2]
    requirements = GRIAdapter(
        backend_root / "data/manifests/gri_requirement_checklist_v3.json"
    ).load_requirements()
    session.add_all(
        [
            AssessmentRecord(
                assessment_id=f"assessment-v3-export-{index:03d}",
                run_id="run-v3-export",
                report_id="report-v3-export",
                standard_id=requirement.standard_id,
                standard_version=requirement.standard_version,
                disclosure_id=requirement.disclosure_id,
                requirement_id=requirement.requirement_id,
                verdict="unknown",
                rationale="待核实",
                missing_items=[],
                model_called=False,
                review_status="needs_manual_review",
            )
            for index, requirement in enumerate(requirements, start=1)
        ]
    )
    session.commit()


async def test_v3_export_covers_all_577_standard_units_without_fake_context_results(
    api_client,
    api_session,
):
    seed_complete_v3_export_scope(api_session)

    response = await api_client.post(
        "/api/reports/report-v3-export/exports/draft",
        json={
            "formats": ["assessment_xlsx", "print_html"],
            "created_by": "产品验收",
        },
    )

    assert response.status_code == 200
    body = response.json()
    scope = body["review_scope"]
    assert scope["standard_unit_total"] == 577
    assert scope["human_reviewed_total"] == 0
    assert "全部 577 项均已人工确认" not in scope["review_scope_statement"]

    _, xlsx_download = await download_export_file(
        api_client,
        body,
        "assessment_xlsx",
    )
    workbook = load_workbook(BytesIO(xlsx_download.content), read_only=True)
    sheet = workbook["GRI核查"]
    assert "AI建议未经人工确认时不构成最终披露结论" in sheet["A1"].value
    headers = [cell.value for cell in sheet[2]]
    data = [
        dict(zip(headers, row))
        for row in sheet.iter_rows(min_row=3, values_only=True)
    ]
    assert len(data) == 577
    assessed = [row for row in data if row["unit_status"] == "assessed"]
    context = [
        row for row in data if row["unit_status"] == "context_incorporated"
    ]
    assert len(assessed) == 499
    assert len(context) == 78
    assert all(row["effective_verdict"] for row in assessed)
    assert all(
        row["effective_verdict"] is None
        and row["review_priority"] is None
        and row["review_status"] is None
        and row["source_pdf_pages"] in (None, "[]")
        for row in context
    )
    workbook.close()

    _, html_download = await download_export_file(
        api_client,
        body,
        "print_html",
    )
    html = html_download.text
    assert "共 577 项" in html
    assert html.count('data-unit-status="assessed"') == 499
    assert html.count('data-unit-status="context_incorporated"') == 78
    assert "已作为上下文纳入相关判断" in html
    assert "全部 577 项均已人工确认" not in html
